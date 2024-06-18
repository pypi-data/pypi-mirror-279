from datetime import datetime, timezone, timedelta
import os
import queue
import threading
import time
import json
import signal
import sys
import traceback
import argparse
import csv
import dbf
import pathlib
from openpyxl import Workbook
import shutil

from concurrent.futures import ThreadPoolExecutor
from pymongo import MongoClient, ASCENDING, DESCENDING

from mongo_client_gateway import MongoClientTradeGateway

from constants import (ENCODING, FILES, Exchange, OrderStatus, OrderType, SecurityType, Side)
from utils import (decode_ims_status,side_to_target_type, decode_exchange_id, encode_ims_side, decode_ordtype, decode_ims_side, decode_ims_contract_type, decode_ims_market, check_charset)

from logger import Logger


try:
    import thread
except ImportError:
    import _thread as thread

LL9 = 1000000000


class ImsServer(MongoClientTradeGateway):
    def __init__(self, config_filename, endtime):
        MongoClientTradeGateway.__init__(self, config_filename, endtime)

        self.load_tg_setting(config_filename)
        self.order_lock = threading.Lock()
        self.trade_lock = threading.Lock()
        self.pos_lock = threading.Lock()
        self.acc_lock = threading.Lock()
        

        self.order_row_set = set()
        self.trade_row_set = set()
        self.pos_row_set = set()
        self.acc_row_set = set()
        self.cancel_row_set = set()

        self.insert_orders = {}
        self.oid_to_ref = {}
        self.sids = []
        self.oid_to_traded = {}
        self.oid_to_traded_money = {}
        
        self.sid_to_req = {}
        self.oid_to_req = {}
        self.oid_to_mid = {}
        self.oid_to_acc = {}
        self.db_id_to_oid = {}
        self.ref_to_oid = {}
        
        self.order_index = -1 #dbf文件当前index
        self.trade_index = -1
        self.oid_to_local_ids = {}
        
    def monitor(self):
        self.thread_pool.submit(self.monitor_algo_order_insert)
        self.thread_pool.submit(self.monitor_cancel_order_insert)
        #self.thread_pool.submit(self.monitor_order_update)

        self.thread_pool.submit(self.monitor_trade_update)

        self.thread_pool.submit(self.date_change)
        #self.thread_pool.submit(self.monitor_pos_update)
        #self.thread_pool.submit(self.monitor_acc_update)
        
    def load_tg_setting(self, config_filename):
        try:
            #config_filename = os.path.join(config_path, 'atx_server_config.json')
            f = open(config_filename, encoding='gbk')
            setting = json.load(f)
            #self.insert_order_path = setting['insert_order_path'].replace('/', '\\')
            self.recv_msg_dir = setting['recv_msg_path'].replace('/', '\\')
            self.insert_order_msg_dir = setting['insert_order_msg_path'].replace('/', '\\')


        except Exception as e:
            err = traceback.format_exc()
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback, limit=None, file=sys.stdout)
            print(f"load config failed! (exception){err}")
            exit(0)

    @MongoClientTradeGateway.error_handler
    def str2bool(self, v:str):
        if isinstance(v, bool):
            return v
        if v.lower() in ('yes', 'true', 't', 'y', '1'):
            return True
        elif v.lower() in ('no', 'false', 'f', 'n', '0'):
            return False
        else:
            raise argparse.ArgumentTypeError('Boolean value expected.')
    
    @MongoClientTradeGateway.error_handler
    def req_buy_order_insert(self, obj):
        workbook = Workbook()

        sheet = workbook.active

        headers = [
            '合约代码',
            '市场',
            '交易类型',
            '投资组合',
            '合约类型',
            '是否市价',
            '指令价格',
            '指令数量',
            '算法策略类型',
            '算法开始时间',
            '算法结束时间',
            '算法参数',
            '备注'
        ]

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        target_account_name = obj['accountName']
        acc = self.target_account_names_to_acc[target_account_name]
        account_id = self.account_id[acc]

        oid = str(self.gen_order_id())
        mid = str(obj['mid'])
        vol =  int(obj["volume"])
        side = int(obj["side"])
        # Write some sample data to the rows below

        target_type = decode_ims_side(side)
        basket_name = mid
        order_vol = f'{vol:d}'
        int_stock_code = int(obj['ticker'])
        stock_code = f'{int_stock_code:06d}'
        market = decode_ims_market(stock_code)
        algo_type = decode_ordtype(obj['executionPlan']["order_type"])

        begin_time = str(obj['executionPlan']["start_time"]).replace(":", "")
        end_time = str(obj['executionPlan']["end_time"]).replace(":", "")

        trade_param = ""
        ExternalId = oid
        #contract_code = obj['contract_code']

        combo_id = obj['combo_id']
        int_contract_type = self.contract_type[acc]
        contract_type = decode_ims_contract_type(int_contract_type)
        is_market = obj['is_market']
        price = int(obj['price'])
        str_limit = ""
        if is_market == 0:
            str_limit = "LIMIT"
        elif is_market == 1:
            str_limit = "ANY"
        
        data = [
            [stock_code, market, target_type, combo_id, contract_type, str_limit, price, vol, algo_type, begin_time, end_time, trade_param, ExternalId],
        ]

        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value

        temp_filename = self.insert_order_msg_dir + '_temp'+ '\\' + 'buy' + oid + '.xlsx'
        # Save the workbook
        workbook.save(temp_filename)

        workbook.close()

        # Move the temporary file to the final filename
        final_filename = self.insert_order_msg_dir + '\\' + 'buy' + oid + '.xlsx'
        shutil.move(temp_filename, final_filename)

        self.logger.info(f"[req_buy_order] insert_success (oid){oid} (mid){mid} (param){data}")
        db_id = obj['_id']
        order_dict = {
            'oid' : oid,
            'mid' : mid,
            "db_id" : db_id,
            "ticker" : obj['ticker'],
            "side" : side,
            "volume" : order_vol,
            "accountName" : target_account_name,
            "order_type" : str(obj['executionPlan']["order_type"]),
            "start_time" : str(obj['executionPlan']["start_time"]),
            "end_time" : str(obj['executionPlan']["end_time"]),
            "contract_type" : int_contract_type,
            "price" : obj['price'],
            "is_market": obj['is_market'],
            "combo_id" : combo_id
        }
        
        target_collection =  self.order_info_db[acc]['target']
        delete_query = {
                    '_id' : db_id
                }
        delete_res = target_collection.delete_one(delete_query)
        self.logger.info(f"[on_req_buy_order] delete (target){obj}")
        
        atx_order_collection = self.order_info_db[acc]['atx_order']
        local_id_list = []
        db_msg = {
                    "oid" : oid,
                    "mid" : mid,
                    "traded_vol" : 0,
                    "traded_amt" : 0,
                    "local_ids" : local_id_list,
                    "order_msg": order_dict
                }
        query = {"oid": oid}
        res = atx_order_collection.replace_one(query, db_msg, True)
        self.logger.info(f"[on_req_order_insert] insert_atx_order_info (res){res} (msg){db_msg}")

        self.sids.append(mid)

        self.sid_to_req[mid] = order_dict
        self.oid_to_req[oid] = order_dict
        self.oid_to_mid[oid] = mid
        self.db_id_to_oid[db_id] = oid
        self.oid_to_acc[oid] = acc
        self.oid_to_traded[oid] = 0
        self.oid_to_traded_money[oid] = 0
        self.oid_to_local_ids[oid] = local_id_list

    @MongoClientTradeGateway.error_handler
    def req_sell_order_insert(self, obj):
        workbook = Workbook()

        sheet = workbook.active

        headers = [
            '合约代码',
            '市场',
            '交易类型',
            '投资组合',
            '合约类型',
            '是否市价',
            '指令价格',
            '指令数量',
            '算法策略类型',
            '算法开始时间',
            '算法结束时间',
            '算法参数',
            '备注'
        ]

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        target_account_name = obj['accountName']
        acc = self.target_account_names_to_acc[target_account_name]
        account_id = self.account_id[acc]

        oid = str(self.gen_order_id())
        mid = str(obj['mid'])
        vol =  int(obj["volume"])
        side = int(obj["side"])
        # Write some sample data to the rows below

        target_type = decode_ims_side(side)
        basket_name = mid
        order_vol = f'{vol:d}'
        int_stock_code = int(obj['ticker'])
        stock_code = f'{int_stock_code:06d}'
        market = decode_ims_market(stock_code)
        algo_type = decode_ordtype(obj['executionPlan']["order_type"])

        begin_time = str(obj['executionPlan']["start_time"]).replace(":", "")
        end_time = str(obj['executionPlan']["end_time"]).replace(":", "")

        trade_param = ""
        ExternalId = oid
        #contract_code = obj['contract_code']

        combo_id = obj['combo_id']
        int_contract_type = self.contract_type[acc]
        contract_type = decode_ims_contract_type(int_contract_type)
        is_market = obj['is_market']
        price = int(obj['price'])
        str_limit = ""
        if is_market == 0:
            str_limit = "LIMIT"
        elif is_market == 1:
            str_limit = "ANY"
        data = [
            [stock_code, market, target_type, combo_id, contract_type, str_limit, price, vol, algo_type, begin_time, end_time, trade_param, ExternalId],
        ]

        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value

        temp_filename = self.insert_order_msg_dir + '_temp' + '\\' + 'sell' + oid + '.xlsx'
        # Save the workbook
        workbook.save(temp_filename)
        
        workbook.close()

        # Move the temporary file to the final filename
        final_filename = self.insert_order_msg_dir + '\\' + 'sell' + oid + '.xlsx'
        shutil.move(temp_filename, final_filename)

        self.logger.info(f"[req_sell_order] insert_success (oid){oid} (mid){mid} (param){data}")
        db_id = obj['_id']
        order_dict = {
            'oid' : oid,
            'mid' : mid,
            "db_id" : db_id,
            "ticker" : obj['ticker'],
            "side" : side,
            "volume" : order_vol,
            "accountName" : target_account_name,
            "order_type" : str(obj['executionPlan']["order_type"]),
            "start_time" : str(obj['executionPlan']["start_time"]),
            "end_time" : str(obj['executionPlan']["end_time"]),
            "contract_type" : int_contract_type,
            "price" : obj['price'],
            "is_market": obj['is_market'],
            "combo_id" : combo_id
        }

        target_collection = self.order_info_db[acc]['sell_target']
        delete_query = {
                    "_id" : db_id
                }
        delete_res = target_collection.delete_one(delete_query)
        self.logger.info(f"[on_req_sell_order] delete (target){obj}")

        atx_order_collection = self.order_info_db[acc]['atx_order']
        local_id_list = []
        db_msg = {
                    "oid" : oid,
                    "mid" : mid,
                    "traded_vol" : 0,
                    "traded_amt" : 0,
                    "local_ids" : local_id_list,
                    "order_msg": order_dict
                }
        query = {"oid": oid}
        res = atx_order_collection.replace_one(query, db_msg, True)
        self.logger.info(f"[on_req_order_insert] insert_sell_order_info (res){res} (msg){db_msg}")

        self.sids.append(mid)
        self.sid_to_req[mid] = order_dict
        self.oid_to_req[oid] = order_dict
        self.oid_to_mid[oid] = mid
        self.db_id_to_oid[db_id] = oid
        self.oid_to_acc[oid] = acc
        self.oid_to_traded[oid] = 0
        self.oid_to_traded_money[oid] = 0
        self.oid_to_local_ids[oid] = local_id_list

    #order_insert 和 rtn_order 都在此获取
    #0 已报 #2 filled 全成 #4 canceled #8 rejected
    @MongoClientTradeGateway.error_handler
    def monitor_order_update(self):
        print ("[monitor_order_update]")
        while not self.is_stopped:
            with self.order_lock:
                order_filename = self.recv_msg_dir  + '\\' + FILES.ORDERS + '.txt'
                                # Specify the path to the input text file

                # Create an empty list to store the dictionaries for each row
                data_list = []

                # Open the file in read mode
                with open(order_filename, 'r') as file:
                    # Read each line in the file
                    for line in file:
                        # Skip lines starting with '#'
                        if line.startswith('#'):
                            continue
                        
                        # Remove leading/trailing whitespaces and split the data using '|'
                        values = line.strip().split('|')
                        
                        # Create a dictionary for the row
                        row_dict = {
                            '产品': values[0],
                            '资产单元': values[1],
                            '投资组合': values[2],
                            '资金帐号': values[3],
                            '市场': values[4],
                            '合约代码': values[5],
                            '序列号': values[6],
                            '委托时间': values[7],
                            '合同号': values[8],
                            '委托数量': values[9],
                            '委托价格': values[10],
                            '买卖方向': values[11],
                            '成交数量': values[12],
                            '成交价格': values[13],
                            '成交编号': values[14],
                            '剩余数量': values[15],
                            '成交状态': values[16],
                            '投资顾问': values[17],
                            '投机套保标志': values[18],
                            '开仓平仓标志': values[19],
                            '篮子ID': values[20],
                            '备注': values[21],
                            '成交时间': values[22],
                            '币种': values[23],
                            '算法ID': values[24],
                            '指令ID': values[25],
                            '审核员': values[26],
                            '交易员': values[27],
                            '错误代码': values[28],
                            '错误消息': values[29],
                            '业务类型': values[30],
                            '子业务类型': values[31],
                            'BookCode': values[32]
                        }
                        
                        # Add the row dictionary to the list
                        data_list.append(row_dict)

                # Print the data list (optional)
                for row_dict in data_list:

                    if row_dict['成交状态'] in ['F_FILLED', 'CANCELED']:
                        self.update_order(row_dict)
                    else:
                        self.logger.info("[monitor_order_update] order_not_finished_or_not_init")


            time.sleep(20)
    
    @MongoClientTradeGateway.error_handler
    def update_order(self,record):
        oid = str(record['备注'])
        acc = ""
        mid = ""
        order_dict = {}
        mudan_id = str(record['投资组合Id'])
        if not oid in self.oid_to_acc: #shut_down or other problem
            trade_acc = str(record['资产单元Id'])

            if trade_acc not in self.account_id_to_acc:
                self.logger.error(f"[update_order] can't_parse_trade_acc {trade_acc}")
                return
            acc = self.account_id_to_acc[trade_acc]
            order_info_collection = self.order_info_db[acc]['atx_order']
            query = {'oid' : oid}
            order_info_target = order_info_collection.find_one(query)
            if not order_info_target is None:
                order_dict = order_info_target['order_msg']
                mid = order_info_target['mid']
                self.oid_to_traded_money[oid] = 0
                self.oid_to_traded[oid] = 0
                self.oid_to_mid[oid] = mid
                self.oid_to_req[oid] = mid
                self.oid_to_ref[oid] = mudan_id
            else:
                self.logger.error(f"[update_order] can't_find_oid (oid){oid}")
                return
        else:
            acc = self.oid_to_acc[oid]
            order_dict = self.oid_to_req[oid]
            mid = self.oid_to_mid[oid]
        ticker = str(record['证券代码'])
        exchange = decode_exchange_id(ticker)
        volume = record['指令数量']
        traded_vol = record['成交数量']
        price = int(record['成交金额']) / int(traded_vol)


        tg_name = order_dict['accountName'].split('@')[0]
        accountName = order_dict['accountName'].split('@')[1]
        order_type = order_dict['order_type']
        start_time = order_dict['start_time']
        end_time = order_dict['end_time']

        target_type = side_to_target_type(order_dict['side'])
        filled_vol = record['成交数量']
        status = decode_ims_status(record['成交状态'])
        
        replace_collection = self.tradelog_db[acc]['order']
        query = {'mid':mid, '_id': int(oid)}
        replace_target = replace_collection.find_one(query)
        if not replace_target is None:

            utc_update_time = replace_target['dbTime']

            self.oid_to_ref[oid] = mudan_id
            order_msg = {
                    "_id": int(oid),
                    "tg_name": tg_name,  # 对应 account_info._id
                    # 对应 tlclient.trader.constant 包 ExchangeID
                    "exchange": exchange,
                    "target_type": target_type,    # 'buy' | 'sell' | 'limit_sell'
                    "volume": volume,  # 订单的volume
                    "price": price,  # 实际成交均价
                    "order_type": order_type,  # algotype:200,目前没有其他的
                    "ticker": ticker,
                    "mid": mid,  # target中对应的 母单sid
                    "accountName": accountName,  # 对应 account_info.account_name
                    "algo_args": {  # 具体的算法参数
                        "order_type": order_type,  # 此柜台不使用下单时要求的算法
                        "start_time": start_time,
                        "end_time": end_time
                    },
                    "status": status,  # 'active' | 'filled' | 'canceled'
                    "filled_vol": filled_vol,  # 实际成交的 volume
                    "dbTime": utc_update_time,
            }
            
            res = replace_collection.replace_one(query, order_msg, True)
            self.logger.info(f"[rtn_order] (res){res} (order_msg){order_msg}")
        else:
            str_update_time = str(record['下单时间'])
            update_time = datetime.strptime(str_update_time, "%Y-%m-%d %H:%M:%S")

            utc_update_time = update_time - timedelta(hours=8)
            #utc_update_time = replace_target['dbTime']

            self.oid_to_ref[oid] = mudan_id
            order_msg = {
                    "_id": int(oid),
                    "tg_name": tg_name,  # 对应 account_info._id
                    # 对应 tlclient.trader.constant 包 ExchangeID
                    "exchange": exchange,
                    "target_type": target_type,    # 'buy' | 'sell' | 'limit_sell'
                    "volume": volume,  # 订单的volume
                    "price": price,  # 实际成交均价
                    "order_type": order_type,  # algotype:200,目前没有其他的
                    "ticker": ticker,
                    "mid": mid,  # target中对应的 母单sid
                    "accountName": accountName,  # 对应 account_info.account_name
                    "algo_args": {  # 具体的算法参数
                        "order_type": order_type,  # 此柜台不使用下单时要求的算法
                        "start_time": start_time,
                        "end_time": end_time
                    },
                    "status": status,  # 'active' | 'filled' | 'canceled'
                    "filled_vol": filled_vol,  # 实际成交的 volume
                    "dbTime": utc_update_time,
            }
            
            res = replace_collection.insert_one(order_msg)
            self.logger.info(f"[rtn_order] (res){res} (order_msg){order_msg}")


    @MongoClientTradeGateway.error_handler
    def monitor_algo_order_insert(self):
        while not self.is_stopped:
            print ("[monitor_algo_order_insert]")
            for acc in self.accounts_run:
                target_account_name = self.target_account_names[acc]
                sell_query = {"accountName": target_account_name}
                sell_targets = self.order_info_db[acc]["short_target"].find(sell_query)
                if sell_targets.count() == 0:
                    self.logger.warning(
                        f"[monitor_sell_order] no_sell_target (acc){acc}")
                    continue
                for sell_target in sell_targets:
                    if sell_target['_id'] not in self.sell_order_db_ids:
                        self.sell_order_db_ids.append(sell_target['_id'])
                        self.req_sell_order_insert(sell_target)
                    else:
                        self.logger.warning(f"[monitor_algo_order_insert] _id_existed (_id){sell_target['_id']}")
            time.sleep(self.scan_interval)
    
    @MongoClientTradeGateway.error_handler
    def monitor_cancel_order_insert(self):
        while not self.is_stopped:
                for acc in self.accounts_run:
                    print ("[monitor_cancel_order_insert]")
                    target_account_name = self.target_account_names[acc]
                    query = {"accountName": target_account_name}
                    with self.cancel_orderlock:
                        cancel_targets = self.order_info_db[acc]["cancel_target"].find(
                            query)
                        if cancel_targets.count() == 0:
                            self.logger.warning(
                                "[monitor_cancel_order] no_cancel_info")
                            continue
                        for cancel_target in cancel_targets:
                            if cancel_target['_id'] not in self.cancel_order_ids:
                                self.cancel_order_ids.append(cancel_target['_id'])
                                self.req_cancel_order(cancel_target)
                time.sleep(self.scan_interval)

    @MongoClientTradeGateway.error_handler
    def req_cancel_order(self, cancel_target):
        oid = str(cancel_target['oid'])
        mudan_id = ""
        if oid not in self.oid_to_ref:
           self.logger.error("[req_cancel_order] can't_find_mudanid (oid){oid}")
           return
        mudan_id = self.oid_to_ref[oid]
        cxltype = 1
        row = [mudan_id, cxltype]
        self.cancel_order_table.append(row)
        
        origin_req = self.oid_to_req[oid]
        target_account_name = origin_req['accountName']
        acc = self.target_account_names_to_acc[target_account_name]
        target_collection =  self.order_info_db[acc]['cancel_target']
        delete_query = {
            'accountName' : target_account_name,
            'oid' : oid
            }
        delete_res = target_collection.delete_one(delete_query)
        self.logger.info(f"[on_req_cancel_order] delete (target){cancel_target}")
    
    #得到的数据 类似于子单的rtn_order
    @MongoClientTradeGateway.error_handler
    def monitor_trade_update(self):
        fields = ["下单时间", "产品名称", "投资组合名称", "产品Id", "资产单元Id", "投资组合Id", "合同号", "交易类型", "发送标志", "市场",
          "证券代码", "证券名称", "成交状态", "买卖方向", "指令数量", "价格类型", "指令价格", "批次号", "订单类型", "指令金额",
          "策略参数", "合约价格", "成交数量", "成交金额", "撤单数量", "录入用户", "备注", "计价币种", "交易汇率", "用户类型",
          "是否DFD", "交易日期", "是否计算佣金", "合约类型"]

        while not self.is_stopped:
            with self.trade_lock:
                trade_filename = self.recv_msg_dir + '\\' + FILES.TRADES  + '.txt'
                data_list = []

                with open(trade_filename, 'r', encoding=check_charset(trade_filename)) as file:

                    # Read each line in the file
                    for line in file:
                        # Skip lines starting with '#'
                        if line.startswith('#'):
                            continue

                        # Remove leading/trailing whitespaces and split the data using '|'
                        values = line.strip().split('|')


                        row_dict = {fields[i]: values[i] for i in range(len(fields))}

                        # Add the row dictionary to the list
                        data_list.append(row_dict)

                        for record in data_list:

                            if record['发送标志'] in ['已接收'] and not record['成交状态'] in ['未成交']: #only need filled/canceled
                                self.on_trade_msg(record)
                                self.update_order(record)
                            elif record['发送标志'] in ['未接收', '已拒绝']:
                                self.logger.info(f"[monitor_trade_update] trade_not_finished_or_rejected msg{record}")


            time.sleep(30)

    @MongoClientTradeGateway.error_handler
    def on_trade_msg(self, record):
        if record['成交数量'] == 0 :
            self.logger.warning(f"[on_trade_msg] nothing_traded (record){record}")
            return
        oid = str(record['备注'])
        mudan_id = str(record['投资组合Id'])

        acc = ""
        mid = ""
        order_dict = {}
        if not oid in self.oid_to_acc: #shut_down or other problem
            trade_acc = str(record['资产单元Id'])
            acc = self.account_id_to_acc[trade_acc]
            order_info_collection = self.order_info_db[acc]['atx_order']
            query = {'oid' : oid}
            order_info_target = order_info_collection.find_one(query)

            if not order_info_target is None:
                order_dict = order_info_target['order_msg']
                mid = order_info_target['mid']
                self.oid_to_local_ids[oid] = order_info_target['local_ids']
                self.oid_to_traded[oid] = 0
                self.oid_to_traded_money[oid] = 0
                self.oid_to_mid[oid] = mid
                self.oid_to_req[oid] = order_dict
                self.oid_to_ref[oid] = mudan_id
            else:
                self.logger.error(f"[on_trade_msg] can't_find_oid (oid){oid}")
                return
        else:
            acc = self.oid_to_acc[oid]
            order_dict = self.oid_to_req[oid]
            mid = self.oid_to_mid[oid]
        _id = str(record['合同号'])
        if _id in self.oid_to_local_ids[oid]:
            #self.logger.warning(f"[on_trade_msg] duplicate_trade_msg (record){record}")
            return
        else:
            self.oid_to_local_ids[oid].append(_id)
        trade_ref = str(record['合同号']) #外部委托编号，出现在这个位置主要是凑数
        ticker = str(record['证券代码'])
        exchange = decode_exchange_id(ticker)
        traded_vol = int(record['成交数量'])
        traded_price = int(record['成交金额']) / traded_vol
        trade_amt = int(record['成交金额'])

        target_type = encode_ims_side(record['买卖方向'])
        if target_type == 1 or target_type == 4:
            self.oid_to_traded_money[oid] += trade_amt
            self.oid_to_traded[oid] += traded_vol
        elif target_type == 2 or target_type == 3:
            self.oid_to_traded_money[oid] -= trade_amt
            self.oid_to_traded[oid] -= traded_vol
        entrust_vol = record['指令数量']
        #entrust_price = record.Price
        entrust_price = float(str(record['指令价格']).strip(' '))
        dbTime = datetime.now(timezone.utc)

        commission = 0
        #str_trade_time = str(int(int(record['成交时间'])/1000))
        #trade_time = datetime.strptime(str_trade_time, "%Y%m%d%H%M%S")
        #utc_trade_time = trade_time - timedelta(hours=8)
        order_type = order_dict['order_type']
        target_account_name = order_dict['accountName']
        acc = self.target_account_names_to_acc[target_account_name]
        log_account_name = order_dict['accountName'].split('@')[1]
        tg_name =  order_dict['accountName'].split('@')[0]

        side = encode_ims_side(record['买卖方向'])

        db_msg = {
                    "trade_ref": _id, # 用的是非凸的子单order_id 经测试之前用的trade_ref 经常为空
                    "oid": oid,
                    "tg_name": tg_name,
                    "exchange": exchange,
                    "ticker": ticker,
                    "traded_vol": traded_vol,
                    "traded_price": traded_price,
                    "order_type": order_type,
                    "side": side,  # 对应 tlclient.trader.constant 包 Side
                    "entrust_vol": entrust_vol,
                    "entrust_price": entrust_price,
                    "dbTime": dbTime,
                    "mid": mid,  # 对应订单中的 mid
                    "commission": 0,  # 没有
                    "trade_time": dbTime,  # 具体交易时间

                    "accountName":  log_account_name,  # 对应 account_info.account_name
                }
        trade_collection = self.tradelog_db[acc]['trade']
        replace_trade_query = { "trade_ref": _id, "oid" : oid, "mid": mid, "accountName": log_account_name}
        db_res = trade_collection.replace_one(replace_trade_query, db_msg, True)
        self.logger.info(
            f"[rtn_trade] (db_res){db_res} (db_msg){db_msg} (traded_vol){traded_vol} (traded_price){traded_price}")
        self.update_position(traded_vol, order_dict, mid, oid, side, trade_amt,exchange)

        atx_order_collection = self.order_info_db[acc]['atx_order']
        atx_order_msg = {
                    "local_ids" : self.oid_to_local_ids[oid],
                }
        update_atx_order_msg = {"$set": atx_order_msg}
        query = {"oid": oid}
        res = atx_order_collection.update_one(query, update_atx_order_msg)
        self.logger.info(f"[on_trade_msg] update_atx_order_info (res){res} (msg){update_atx_order_msg}")

    @MongoClientTradeGateway.error_handler
    def update_position(self, traded_vol, order, mid , oid, side, trade_amt, exchange):
        target_account_name = order['accountName']
        acc = self.target_account_names_to_acc[target_account_name]

        query = {'mid': mid, 'accountName': target_account_name}
        position_collection = self.order_info_db[acc]['EquityPosition']
        position = position_collection.find_one(query)
        if position is None and traded_vol > 0:
            if oid not in self.oid_to_req:
                self.logger.warning(
                    f"[update_position] can't_find_req (mid){mid} (oid){oid}")
                return
            if oid not in self.oid_to_traded:
                self.logger.error(f"[update_position]no trade_vol (mid){mid}")
                return
            td_trade_vol = self.oid_to_traded[oid]

            amt = self.oid_to_traded_money[oid]

            ticker = order['ticker']
            yd_pos = 0
            mkt_value = amt
            cost = 0
            if not td_trade_vol == 0:
                cost = amt / td_trade_vol
            update_time = datetime.now(timezone.utc)

            position_msg = {
                "ticker": ticker,
                "cost": cost,  # 成本价
                "td_pos_long": td_trade_vol,  # 今仓
                "yd_pos_long": yd_pos,  # 昨仓（可卖），挂止盈单后为0
                "td_pos_short": 0,
                "yd_pos_short": 0,
                "actual_td_pos_long": td_trade_vol,  # 实际的今仓
                "actual_yd_pos_long": yd_pos,  # 实际的昨仓
                "enter_date": update_time,
                "holding_days": 0,  # 持仓时间（交易日累计）
                "mkt_value": mkt_value,  # 市值
                "exchange": exchange,# 对应 tlclient.trader.constant 包 ExchangeID
                "mid": mid,  # 我们策略生成的id，需要根据实际下单情况维护各sid的持仓。
                # 格式: {account_info.tg_name}@{account_info.account_name}
                "accountName": target_account_name,
                "update_date": update_time,
            }
            query2 = {'mid': mid,
                        'accountName': target_account_name}
            res = position_collection.replace_one(query2, position_msg, True)
            self.logger.info(
                f"[update_position] (res){res} (msg){position_msg}")
        else:
            if oid not in self.oid_to_req:
                self.logger.warning(
                    f"[update_position] can't_find_req (mid){mid}")
                return
            if oid not in self.oid_to_traded:
                self.logger.error(f"[update_position]no trade_vol (mid){mid}")
                return
            actual_yd_pos = position['actual_yd_pos_long']
            amt = trade_amt
            yd_pos = position['yd_pos_long']
            if side == 1 or side == 3:
                td_trade_vol = position['actual_td_pos_long'] + traded_vol
                mkt_value = position['mkt_value'] + amt
            elif side == 2 :
                td_trade_vol = position['actual_td_pos_long'] - traded_vol
                mkt_value = position['mkt_value'] - amt
                yd_pos = yd_pos - traded_vol

            ticker = order['ticker']

            cost = 0
            total_vol = td_trade_vol + yd_pos
            if not total_vol == 0:
                cost = mkt_value / total_vol
            update_time = datetime.now(timezone.utc)
            accountName = order['accountName']
            enter_time = position['enter_date']
            holding_days = position['holding_days']
            position_msg = {
                "ticker": ticker,
                "cost": cost,  # 成本价
                "td_pos_long": td_trade_vol,  # 今仓
                "yd_pos_long": yd_pos,  # 昨仓（可卖），挂止盈单后为0
                "td_pos_short": 0,
                "yd_pos_short": 0,
                "actual_td_pos_long": td_trade_vol,  # 实际的今仓
                "actual_yd_pos_long": actual_yd_pos,  # 实际的昨仓
                "enter_date": enter_time,
                "holding_days": holding_days,  # 持仓时间（交易日累计）
                "mkt_value": mkt_value,  # 市值
                "exchange": exchange,  # 对应 tlclient.trader.constant 包 ExchangeID
                "mid": mid,  # 我们策略生成的id，需要根据实际下单情况维护各sid的持仓。
                # 格式: {account_info.tg_name}@{account_info.account_name}
                "accountName": accountName,
                "update_date": update_time,
            }
            query = {'mid': mid,
                        'accountName': accountName}
            res = position_collection.replace_one(query, position_msg, True)
            self.logger.info(
                f"[update_position] (res){res} (msg){position_msg}")

    def start(self):
        signal.signal(signal.SIGTERM, self.signal_handler)
        signal.signal(signal.SIGINT, self.signal_handler)
        _msg = f"[login] ims_server_start (time){datetime.now()}"
        self.send_to_user(logger=self.logger, url_list=self.url_list, msg=_msg)
        self.monitor()

    def close(self):
        return super().close()
    
    def join(self):
        while self.is_stopped == False:
            time.sleep(0.01)
            if self.is_stopped:
                self.logger.info(
                    "[close] main thread is stopped,active_orders message will lose")
                self.close()

if __name__ == "__main__":

    sys.stdout.reconfigure(encoding='utf-8')
    description = "atx_server,get target from mongodb and serve atx"  
    parser = argparse.ArgumentParser(description=description)

    parser.add_argument('-e' , '--end_time', dest='end_time', default='15:30')
    _config_filename = "C:/Users/Administrator/Desktop/ims_batandconfig/ims_server_config.json"
    parser.add_argument('-p', '--config_filepath', dest= 'config_filepath', default= _config_filename)

    args = parser.parse_args()
    print (f"(args){args}")

    td = ImsServer(args.config_filepath, args.end_time)
    td.start()

    td.join()