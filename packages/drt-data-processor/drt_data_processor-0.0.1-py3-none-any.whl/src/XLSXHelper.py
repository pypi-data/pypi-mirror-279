import errno
import math
import openpyxl
import pandas
from openpyxl import Workbook


def read_csv(filename):
    data = pandas.read_csv(filename, sep=';')
    return data


def read_excel(filename: str) -> pandas.DataFrame:
    data = pandas.read_excel(filename)
    return data


def calculate_distance(lat1, lon1, lat2, lon2):
    r = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2) * math.sin(dlat/2) + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2) * math.sin(dlon/2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    distance = r * c
    return distance


def find_distances(d):
    distances = []
    for i in range(0, d.shape[0]):
        point1 = d.at[i, 'Координата выпуска'].split(',')
        point2 = d.at[i, 'Координата ниже'].split(',')
        distance = calculate_distance(float(point1[0]), float(point1[1]),
        float(point2[0]), float(point1[1])) * 1000
        distances.append(distance)

    d.insert(5, "Расстояние, м", distances, True)
    return d


def write_excel(filename: str, arr: list, x: float, step: int, deltaZ: float, count=1) -> int:
    try:
        book = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(errno)
        book = Workbook()

    sheet = book.active

    row_count = count

    if step == 0:
        sheet.cell(row=1, column=1).value = \
            f"Kарта распространения вещества: x = 0, step = 0, deltaZ = deltaY = {round(deltaZ, 2)}"
    else:
        sheet.cell(row=row_count, column=1).value = f'x = {x}, step = {step}'

    row_count += 1
    for row in arr:
        sheet.append(row)
        row_count += 1

    book.save(filename)
    return row_count
