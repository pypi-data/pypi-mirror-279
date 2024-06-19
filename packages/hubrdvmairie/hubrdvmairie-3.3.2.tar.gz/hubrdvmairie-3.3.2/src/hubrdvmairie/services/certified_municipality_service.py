import io
import json
import logging
from collections import defaultdict
from typing import List

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from src.hubrdvmairie.models.certified_municipality import CertifiedMunicipality

from ..crud.crud_certified_municipality import certified_municipality as crud

logger = logging.getLogger(__name__)


def get_all(session) -> List[CertifiedMunicipality]:
    try:
        return crud.get_all(session)
    except Exception as e:
        logger.error("Error while getting all certified_municipality : %s", str(e))


def get_all_certified_municipalities(
    session: sessionmaker,
) -> List[CertifiedMunicipality]:
    municipalities = get_all(session)
    certified_municipalities_complete = []
    for certified_municipality in municipalities:
        complete = certified_municipality.complete()
        certified_municipalities_complete.append(complete)
    departments = defaultdict(list)

    for municipality in certified_municipalities_complete:
        department_code = municipality["department_code"]
        departments[department_code].append(municipality)

    departments_list = []
    for department_code, municipalities in departments.items():
        department_info = {
            "department_code": department_code,
            "department_name": municipalities[0]["department_name"],
            "municipalities": municipalities,
        }
        departments_list.append(department_info)
    departments_list = sorted(departments_list, key=lambda x: x["department_code"])
    return departments_list


async def update_certified_municipality_table(
    session: sessionmaker, uploaded_file: UploadFile
):
    certified_municipalities = await read_certified_municipality_from_file_streaming(
        uploaded_file
    )
    create_list = []
    unchanged_list = []
    nb_certified_municipalities = len(certified_municipalities)
    for certified_municipality in certified_municipalities:
        res = crud.save_or_update(session, obj_in=certified_municipality)
        if res[0] == "created":
            create_list.append(res[1])
        else:
            unchanged_list.append(res[1])

    yield json.dumps(
        {
            "nb_certified_municipality": nb_certified_municipalities,
            "created : ": str(len(create_list)),
            "unchanged : ": str(len(unchanged_list)),
        }
    )


async def read_certified_municipality_from_file_streaming(
    uploaded_file: UploadFile,
) -> CertifiedMunicipality:
    # read file depending on its type
    if uploaded_file.filename.endswith(".xlsx"):
        return await read_certified_municipality_file_streaming(uploaded_file)
    else:
        raise TypeError("Unknown file type : " + str(uploaded_file.filename))


async def read_certified_municipality_file_streaming(uploaded_file: UploadFile):
    certified_municipalities = set()

    file_content = await uploaded_file.read()
    xls_data = io.BytesIO(file_content)
    data = load_workbook(xls_data, read_only=True).active
    for ligne in data.iter_rows(min_row=2, max_row=data.max_row, values_only=True):
        ugf = str(ligne[1])
        town_hall_name = ligne[0]
        address = ligne[2]
        zip_code = str(ligne[3])
        city_name = ligne[4]
        phone_number = ligne[5]
        website = ligne[6]
        appointment_details = str(ligne[7])
        service_opening_date = ligne[8]
        label = ligne[9]

        certified_municipality = CertifiedMunicipality(
            ugf=ugf,
            town_hall_name=town_hall_name,
            address=address,
            zip_code=zip_code,
            city_name=city_name,
            phone_number=phone_number,
            website=website,
            appointment_details=appointment_details,
            service_opening_date=service_opening_date,
            label=label,
        )
        certified_municipalities.add(certified_municipality)
    return certified_municipalities
