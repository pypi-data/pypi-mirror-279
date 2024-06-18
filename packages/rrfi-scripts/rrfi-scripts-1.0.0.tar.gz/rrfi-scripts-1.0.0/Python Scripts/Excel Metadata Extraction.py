import os
import uuid
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
import openpyxl
from io import BytesIO
from azure.search.documents import SearchClient
from azure.core.credentials import AzureKeyCredential
import re
import csv
import logging
from openai import AzureOpenAI
import hashlib
import json
 
# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
 
# Azure Blob Storage connection string and container name
connection_string = "DefaultEndpointsProtocol=https;AccountName=rrfimainstorage;AccountKey=9h2WMr2hvNI1V8xixizn49neFq6/Oba61Z4e6L9YAuuRqrSXEsyEB2NuDhO2NXZaBec5VvH4O8D2+ASt1tyDqA==;EndpointSuffix=core.windows.net"
container_name = "testexcel"
 
# Azure Cognitive Search configuration
search_service_name = "rrfimainsearch"
index_name = "rrfiexcelsheets"
api_key = "1LnrhfcdFVeuJ7jVOe2ElJi2uFib34RZhLWnig1nDiAzSeAIlOEO"
search_endpoint = 'https://rrfimainsearch.search.windows.net'
 
# Azure OpenAI configuration
openai_endpoint = "https://rrfimainoai.openai.azure.com/"
openai_api_key = "9c7a9a0b6b584e2c8701ae2fbb0ff6c2"
 
# Connect to Blob Storage
blob_service_client = BlobServiceClient.from_connection_string(connection_string)
container_client = blob_service_client.get_container_client(container_name)
 
# Connect to Azure Cognitive Search
search_client = SearchClient(endpoint=search_endpoint,
                             index_name=index_name,
                             credential=AzureKeyCredential(api_key))
 
# Connect to Azure OpenAI
openai_client = AzureOpenAI(
    azure_endpoint=openai_endpoint,
    api_key=openai_api_key,
    api_version="2024-02-15-preview"
)
 
# Area codes list
area_codes = ['NEA', 'CLA', 'YOR', 'GMC', 'EMD', 'LNA', 'WMD', 'EAN', 'HNL', 'THM', 'KSL', 'SSD', 'WSX', 'DCS']
 
# Regex pattern for Postcode extraction
#postcode_regex = re.compile(r'^([A-Za-z]{2}[\d]{1,2}[A-Za-z]?)[\s]+([\d][A-Za-z]{2})$')
postcode_regex = re.compile(r'^([A-Za-z]{1,2}\d[A-Za-z\d]?)[\s]?(\d[A-Za-z]{2})$')
 
# Load river names from Excel file
def load_river_names_from_excel(file_path):
    river_names = set()  # Use a set to store unique river names
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            river_names.add(row[0])  # Add river name to set
    return list(river_names)  # Convert set back to list before returning
 
# Load river names from Excel file
excel_file_path = r'C:\Users\2049811\OneDrive - Cognizant\Desktop\DEFRA\Uk Rivers.xlsx'
river_names = load_river_names_from_excel(excel_file_path)
 
# Function to generate hash value for a file
def generate_file_hash(blob_content):
    hasher = hashlib.md5()
    hasher.update(blob_content)
    return hasher.hexdigest()
 
# Recursive function to iterate through blobs
def process_blobs(container_client, depth=0):
    # Check for excessive depth to prevent infinite recursion
    if depth > 50:
        logging.warning("Excessive depth detected, stopping recursion.")
        return
 
    # Initialize sets to store unique source names and types
    unique_source_names = set()
    unique_source_types = set()
    unique_property_types = set()
    unique_ta_codes = set()
    unique_ta_names = set()
    unique_ngr = set()
 
    document = {}
 
    # Iterate over blobs in the current container
    for blob in container_client.list_blobs():
        logging.info(f"Processing file: {blob.name}")
        #if '.azDownload' not in blob.name and blob.name.lower().endswith(('.xlsx', '.xls','.xlsm'))
        # If blob is a directory, recursively process blobs in that directory
        #if blob.name.startswith("~$"):
         #   logging.info("skipping temp excel file")
          #  continue
       
        if blob.name[-1] == '/':
            sub_folder_client = blob_service_client.get_container_client(container_name + '/' + blob.name)
            process_blobs(sub_folder_client, depth=depth + 1)
        else:
            try:
               
                # Extract location from the file path for Incident Name
                location_completion = openai_client.chat.completions.create(
                    model="gpt-35-turbo",
                    messages=[
                        {"role": "system", "content": blob.name},
                        {"role": "system", "content": "You are an AI assistant that helps process data in a pipeline. Just return the values without any additional response text. Always return the last location name."},    
                        {"role": "user", "content": "You are a helpful agent. Your task is to identify and extract UK location names as a substring from the given text. It is possible to have multiple location names within the URL text."}
                    ],
                    temperature=0.7,
                    max_tokens=50,
                    top_p=0.95,
                    frequency_penalty=0,
                    presence_penalty=0,
                    stop=None
                )
               
                location = location_completion.choices[0].message.content
 
                # Call OpenAI GPT-3.5 model to generate Incident_Date_and_Time
                Incident_Datetime_completion = openai_client.chat.completions.create(
                    model="gpt-35-turbo",  # model = "deployment_name"
                    messages=[
                        {"role": "system", "content": blob.name},
                        {"role": "system", "content": "You are an AI assistant that helps process data in a pipeline. Just return the values without any additional response text. Always return the identified date in dd/mm/yyyy format. Only return one value based on prioritization. Higher priority value gets preference."},    
                        {"role": "user", "content": "You are a helpful agent. Your task is to identify and extract date, month, year as a substring from the given text. It is possible to have multiple date, month within the URL text. Highest priority for a full date. If there is no full date, look for a month and year value to return."}
                    ],
                    temperature=0.7,
                    max_tokens=800,
                    top_p=0.95,
                    frequency_penalty=0,
                    presence_penalty=0,
                    stop=None
                )
 
                # Extract Incident date and time from OpenAI output
                incident_datetime = Incident_Datetime_completion.choices[0].message.content
 
                #Extracting event type for Incident Name:
                if 'storm' in blob.name.lower() or 'flood' in blob.name.lower():                  
                        path_folders = blob.name.split("/")
                        for event in path_folders:
                            if 'storm' in event.lower() or 'flood' in event.lower():
                                event_folder = event
                                break
                else:
                        event_folder=''
                                             
                # Extract the Area_Name from blob name
                area_name = None
                for code in area_codes:
                    if code in blob.name:
                        area_name = code
                        break
 
                # Extract Incident_Name from full_file_path
                       
                if 'EA-IMToolbox' in blob.name:
                    if 'Current Incidents' in blob.name:
                        path_folders = blob.name.split("/")
                        current_incidents_index = path_folders.index("Current Incidents")
                        IncName = path_folders[current_incidents_index + 1]
                    elif 'Past Incidents' in blob.name:
                        path_folders = blob.name.split("/")
                        current_incidents_index = path_folders.index("Past Incidents")
                        IncName = path_folders[current_incidents_index + 1]
                    else:
                        area_name = area_name or ''
                        event_folder = event_folder or ''
                        location = location or ''
                        incident_datetime = incident_datetime or ''
                        if area_name=='':
                            IncName=event_folder+'_'+location+'_'+incident_datetime
                        elif event_folder=='':
                            IncName=area_name+'_'+location+'_'+incident_datetime
                        elif location=='':
                            IncName=area_name+'_'+event_folder+'_'+incident_datetime
                        elif incident_datetime=='':
                            IncName=area_name+'_'+event_folder+'_'+location
                        else:
                            IncName=area_name+'_'+event_folder+'_'+location+'_'+incident_datetime
                else:
                    #IncName = f"{area_name}_{location}_{event_folder}_{incident_datetime}"
                    area_name = area_name or ''
                    event_folder = event_folder or ''
                    location = location or ''
                    incident_datetime = incident_datetime or ''
                    if area_name=='':
                        IncName=event_folder+'_'+location+'_'+incident_datetime
                    elif event_folder=='':
                        IncName=area_name+'_'+location+'_'+incident_datetime
                    elif location=='':
                        IncName=area_name+'_'+event_folder+'_'+incident_datetime
                    elif incident_datetime=='':
                        IncName=area_name+'_'+event_folder+'_'+location
                    else:
                        IncName=area_name+'_'+event_folder+'_'+location+'_'+incident_datetime
 
                # Check if filename contains 'Flooded Properties'
                properties_flooded = 'Yes' if 'flooded properties' in blob.name.lower()  or 'property flooding' in blob.name.lower() else 'No'
 
                # Check if filename contains 'Flooded Properties'
                property_protected = 'Yes' if 'properties protected' in blob.name.lower() else 'No'
 
                # Extract metadata and send to Azure Cognitive Search
                if area_name:
                    if blob.name.lower().endswith(('.xlsx', '.xls', '.csv', '.xlsm')):
                        logging.info(f"Processing file format: {blob.name}")
                        # Get the blob client
                        blob_client = container_client.get_blob_client(blob)
                        # Download blob data
                        blob_data = blob_client.download_blob()
                        blob_content = blob_data.read()
 
                        # Generate hash value for file
                        hash_value = generate_file_hash(blob_content)
 
                        # Initialize a set to store unique postcodes
                        unique_postcodes = set()
                        # Initialize a set to store unique found river names
                        found_river_names = set()
 
                        # Handle Excel files
                        if blob.name.lower().endswith(('.xlsx', '.xls')):
                            logging.info(f"Processing Excel file: {blob.name}")
                            workbook = load_workbook(filename=BytesIO(blob_content))
                            for sheet in workbook.sheetnames:
                                logging.info(f"Processing sheet: {sheet}")
                                worksheet = workbook[sheet]
 
                                # Convert worksheet content to text for OpenAI input
                                worksheet_text = "\n".join(" ".join(str(cell) for cell in row if cell) for row in worksheet.iter_rows(values_only=True))
 
                                # Call OpenAI GPT-3.5 model to generate location of incident
                                completion = openai_client.chat.completions.create(
                                    model="gpt-35-turbo",  # model = "deployment_name"
                                    messages=[
                                        {"role": "system", "content": worksheet_text},
                                        {"role": "user", "content": "From the given data identify all town names. \nReturn all town names concatenated in single string comma separated. \nOnly return valid town name without any numerical value or special character"},
                                        {"role": "system", "content": "You help process data within a pipeline.\nOnly return the values without any other response text."}
                                    ],
                                    temperature=0.7,
                                    max_tokens=800,
                                    top_p=0.95,
                                    frequency_penalty=0,
                                    presence_penalty=0,
                                    stop=None
                                )
 
                                # Extract location of incident from OpenAI output
                                location_of_incident = completion.choices[0].message.content
 
                                # Update the document with Location_of_Incident field
                                #document['Location_of_Incident'] = location_of_incident
 
                                # Call OpenAI GPT-3.5 model to generate catchment name
                                catchment_completion = openai_client.chat.completions.create(
                                    model="gpt-35-turbo",  # model = "deployment_name"
                                    messages=[
                                        {"role": "system", "content": worksheet_text},
                                        {"role": "user", "content": "From the given data identify all catchment names. \nReturn all catchment names concatenated in a single string comma separated. \nOnly return valid catchment names without any numerical values or special characters"},
                                        {"role": "system", "content": "You help process data within a pipeline.\nOnly return the values without any other response text."}
                                    ],
                                    temperature=0.7,
                                    max_tokens=800,
                                    top_p=0.95,
                                    frequency_penalty=0,
                                    presence_penalty=0,
                                    stop=None
                                )
 
                                # Extract catchment name from OpenAI output
                                catchment_name = catchment_completion.choices[0].message.content
 
                                # Update the document with Catchment_Name field
                                #document['Catchment_Name'] = catchment_name
 
                                # Call OpenAI GPT-3.5 model to generate Security Classification
                                security_completion = openai_client.chat.completions.create(
                                    model="gpt-35-turbo",  # model = "deployment_name"
                                    messages=[
                                        {"role": "system", "content": worksheet_text},
                                        {"role": "user", "content": "From the data you are given, identify if 'Official', 'Secret', 'Top Secret', 'Official Sensitive' is included within the data as text. If any of them are listed within the data, return the listed text."},
                                        {"role": "system", "content": "You help process data in a pipeline. \nOnly return the values without any other response text. \nOnly return the listed text that exist within the data and no other text"}
                                    ],
                                    temperature=0.7,
                                    max_tokens=800,
                                    top_p=0.95,
                                    frequency_penalty=0,
                                    presence_penalty=0,
                                    stop=None
                                )
 
                                # Extract security classification from OpenAI output
                                security_classification = security_completion.choices[0].message.content
 
                                # Update the document with Security_Classification field
                                #document['Security_Classification'] = security_classification
 
                                # Update the document with Security_Classification field
                                #document['Incident_Date_and_Time'] = incident_datetime
 
                                # Upload document to Azure Cognitive Search
                                #search_client.upload_documents(documents=[document])
                                #logging.info("Uploaded document to Azure Cognitive Search.")
 
                                # Extract postcodes from Excel sheet
                                for row in worksheet.iter_rows(values_only=True):
                                    for cell in row:
                                        if cell and postcode_regex.match(str(cell)):
                                            unique_postcodes.add(cell)
                                        # Search for river names
                                        for river in river_names:
                                            if river.lower() in str(cell).lower():
                                                found_river_names.add(river)
                                   
                                                                    
                                    # Extract unique source names from the worksheet
                                    source_name_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and ('source name' in str(header.value).lower() or 'source' in str(header.value).lower() or 'watercourse' in str(header.value).lower() or 'waterbody type' in str(header.value).lower()):
                                            source_name_column_index = i
                                            break
                                    if source_name_column_index is not None:
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            if row is not None:
                                                source_name = row[source_name_column_index]
                                                if source_name and source_name.strip().lower !='unknown':
                                                    unique_source_names.add(source_name.strip())
                                    
                                    
                                    # Extract unique source types from the worksheet
                                    source_type_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and ('source type' in str(header.value).lower() or 'primary source of flooding' in str(header.value).lower() or 'source of flooding' in str(header.value).lower()):
                                            source_type_column_index = i
                                            break
                                    if source_type_column_index is not None:
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            if row is not None:
                                                source_type = row[source_type_column_index]
                                                if source_type and source_type.strip().lower != 'unknown':
                                                    unique_source_types.add(source_type.strip())
                                                                      
                                    
                                    # Extract unique property types from the worksheet
                                    property_type_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and 'property type' in str(header.value).lower():
                                            property_type_column_index = i
                                            break
                                    if property_type_column_index is not None:
                                        property_type = row[property_type_column_index]
                                        if property_type:
                                            unique_property_types.add(property_type)
                                    
                                    '''
                                    # Extract unique property types from the worksheet
                                    property_type_column_index = None
                                    
                                    # Identify the column index for property type
                                    for i, header in enumerate(worksheet[1]):
                                        header_value = str(header.value).strip().lower() if header.value else ''
                                        if 'property type' in header_value:
                                            property_type_column_index = i
                                            break
                                    
                                    # Initialize a set to store unique property types
                                    unique_property_types = set()
                                    
                                    # If the property type column was found, extract unique property types
                                    if property_type_column_index is not None:
                                        for row in worksheet.iter_rows(min_row=2, values_only=True):
                                            if row is not None:
                                                property_type = row[property_type_column_index]
                                                if property_type and property_type.strip().lower() != 'unknown' or property_type and property_type.strip().lower() != 'Unknown':
                                                    unique_property_types.add(property_type.strip())

                                    # Convert set to a list
                                    unique_property_types_list = list(unique_property_types)

                                    # Serialise the list to JSON
                                    unique_property_types_json = json.dumps(unique_property_types_list)
                                    '''

                                    # Extract unique TA code from the worksheet
                                    ta_code_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and ('ta code' in str(header.value).lower() or 'target area code' in str(header.value).lower()):
                                            ta_code_column_index = i
                                            break
                                    if ta_code_column_index is not None:
                                        ta_code = row[ta_code_column_index]
                                        if ta_code:
                                            unique_ta_codes.add(ta_code)
 
                                    # Extract unique TA name from the worksheet
                                    ta_name_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and ('ta name' in str(header.value).lower() or 'target area name' in str(header.value).lower() or 'area name' in str(header.value).lower()):
                                            ta_name_column_index = i
                                            break
                                    if ta_name_column_index is not None:
                                        ta_name = row[ta_name_column_index]
                                        if ta_name:
                                            unique_ta_names.add(ta_code)
 
                                    # Extract unique National Grid from the worksheet
                                    ngr_column_index = None
                                    for i, header in enumerate(worksheet[1]):
                                        if header.value and ('ngr' in str(header.value).lower() or 'national grid' in str(header.value).lower()):
                                            ngr_column_index = i
                                            break
                                    if ngr_column_index is not None:
                                        ngr = row[ngr_column_index]
                                        if ngr:
                                            unique_ngr.add(ngr)
 
                        # Handle CSV files
                        elif blob.name.lower().endswith('.csv'):
                            logging.info(f"Processing CSV file: {blob.name}")
                            csv_content = blob_content.decode('utf-8')
                            csv_reader = csv.reader(csv_content.splitlines())
                            for row in csv_reader:
                                for cell in row:
                                    if cell and postcode_regex.match(str(cell)):
                                        unique_postcodes.add(cell)
                                    # Search for river names
                                    for river in river_names:
                                        if river.lower() in str(cell).lower():
 
                                    # Extract unique source names from the CSV
                                            source_name_index = None
                                    for i, column in enumerate(row):
                                        if column and ('source name' in str(header.value).lower() or 'source' in str(header.value).lower() or 'watercourse' in str(header.value).lower() or 'waterbody type' in str(header.value).lower()):
                                            source_name_index = i
                                            break
                                    if source_name_index is not None:
                                        source_name = row[source_name_index]
                                        if source_name:
                                            unique_source_names.add(source_name)
 
                                    # Extract unique source types from the CSV
                                    source_type_index = None
                                    for i, column in enumerate(row):
                                        if column and ('source type' in str(header.value).lower() or 'primary source of flooding' in str(header.value).lower() or 'source of flooding' in str(header.value).lower()):
                                            source_type_index = i
                                            break
                                    if source_type_index is not None:
                                        source_type = row[source_type_index]
                                        if source_type:
                                            unique_source_types.add(source_type)
 
                                    # Extract unique property types from the CSV
                                    property_type_index = None
                                    for i, column in enumerate(row):
                                        if column and 'property type' in str(column).lower():
                                            property_type_index = i
                                            break
                                    if property_type_index is not None:
                                        property_type = row[property_type_index]
                                        if property_type:
                                            unique_property_types.add(property_type)
 
                                    # Extract unique TA Name from the CSV
                                    ta_name_index = None
                                    for i, column in enumerate(row):
                                        if column and ('ta name' in str(header.value).lower() or 'target area name' in str(header.value).lower() or 'area name' in str(header.value).lower()):
                                            ta_name_index = i
                                            break
                                    if ta_name_index is not None:
                                        ta_name = row[ta_name_index]
                                        if ta_name:
                                            unique_ta_names.add(ta_name)
 
                                    # Extract unique TA Code from the CSV
                                    ta_code_index = None
                                    for i, column in enumerate(row):
                                        if column and ('ta code' in str(header.value).lower() or 'target area code' in str(header.value).lower()):
                                            ta_code_index = i
                                            break
                                    if ta_code_index is not None:
                                        ta_code = row[ta_code_index]
                                        if ta_code:
                                            unique_ta_codes.add(ta_code)
 
                                    # Extract unique National Grid from the CSV
                                    ngr_index = None
                                    for i, column in enumerate(row):
                                        if column and ('ta code' in str(header.value).lower() or 'target area code' in str(header.value).lower()):
                                            ngr_index = i
                                            break
                                    if ngr_index is not None:
                                        ngr = row[ngr_index]
                                        if ngr:
                                            unique_ngr.add(ngr)
 
                        # Handle other file formats, if needed
                        else:
                            logging.warning(f"Skipping unsupported file format: {blob.name}")
 
                        # Generate unique alphanumeric id
                        unique_id = str(uuid.uuid4())
 
                        # Create a document with concatenated postcodes, found river names, and other relevant fields
                        postcodes_str = ', '.join(str(postcode) for postcode in unique_postcodes)
                        river_names_str = ', '.join(found_river_names)
                        document = {
                            "Area_Name": area_name,
                            "File_Name": os.path.basename(blob.name),
                            "File_Path": os.path.dirname(blob.name),
                            "Source_Path": blob.name,
                            "Item_Type": os.path.splitext(os.path.basename(blob.name))[1] if '.' in blob.name else '',
                            "Postcode": postcodes_str,
                            "River_Name": river_names_str,
                            "Incident_Name": IncName,
                            "Properties_Flooded": properties_flooded,
                            "Source_Name": ", ".join(unique_source_names),
                            #"Source_Name": unique_source_names_json,
                            "Source_Type": ", ".join(unique_source_types),
                            #"Source_Type": unique_source_types_json,
                            "Property_Protected": property_protected,
                            "Property_Type": ", ".join(unique_property_types),
                            #"Property_Type": unique_property_types_json,
                            "TA_Name": ", ".join(unique_ta_names),
                            "TA_Code": ", ".join(unique_ta_codes),
                            "National_Grid": ", ".join(unique_ngr),
                            "Location_of_Incident": location_of_incident,
                            "Catchment_Name": catchment_name,
                            "Security_Classification": security_classification,
                            "Incident_Date_and_Time": incident_datetime,
                            "Hash_Value": hash_value,
                            "id": unique_id
                        }
                        search_client.upload_documents(documents=[document])
                        logging.info("Uploaded document to Azure Cognitive Search.")
 
                    else:
                        logging.warning(f"Skipping unsupported file format: {blob.name}")
                else:
                    logging.warning(f"No area code found in file: {blob.name}")
            except Exception as e:
                logging.error(f"An error occurred processing blob {blob.name}: {str(e)}")
 
# Start recursive processing
process_blobs(container_client)